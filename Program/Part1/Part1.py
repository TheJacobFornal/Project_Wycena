import datetime
from openpyxl import load_workbook
from pathlib import Path
import re
from Part2 import Part2 as part_2

curr_dir = Path(__file__).parent
main_dir = curr_dir.parent.parent
excel_dir = main_dir / "Excel"

orders_path = excel_dir / "Zamówienia Kopia JakubF.xlsx"

wb = load_workbook(orders_path)

ws = wb.active

def get_month_sheet():
    months = [
        "Styczeń", "Luty", "Marzec", "Kwiecień", "Maj", "Czerwiec",
        "Lipiec", "Sierpień", "Wrzesień", "Październik", "Listopad", "Grudzień"
    ]

    month_sheet = [name for name in wb.sheetnames if any(name.startswith(m) for m in months)]

    return month_sheet

Excel_map_1 = {
    5: 6,
    6: 7,
    7: 21,
    8: 36,
    9: 25,
    10: 28,
    11: 32,
    12: 34,
    13: 30,
    14: 38,

    18: 8,
    19: 11,
    20: 12,
    21: 13,
    22: 14,
    23: 15
}

Excel_map_2 = {
    5: 6,
    6: 7,
    7: 21,
    8: 36,
    9: 25,
    10: 28,
    11: 32,
    12: 34,
    13: 30,
    14 : 100,                                   ## jaki numer ma spawanie
    15: 38,

    19: 8,
    20: 11,
    21: 12,
    22: 13,
    23: 14,
    24: 15
}

task_table = []

def newer_Excel(ws):
    name = ws.title
    match = re.search(r'\d+', name)

    if match:
        number = int(match.group())
        if number >= 25:
            return True
        else:
            return False
    return None

def add_to_table(ws, row, col):
    value = ws.cell(row, col).value

    if newer_Excel(ws):
        elem_col = Excel_map_2[col]
    else:
        elem_col = Excel_map_1[col]

    task_table.append([elem_col, value])


def look_for_operation(ws, row):                                #go through cols in row
    if newer_Excel(ws):
        for col in range(7, 16):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value is not None:
                add_to_table(ws, row, col)
    else:
        for col in range(7, 15):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value is not None:
                add_to_table(ws, row, col)


def check_if_canceled(ws, row):
    cell_value = ws.cell(row=row, column=1).value
    if isinstance(cell_value, datetime.datetime):
        return False
    else:
        return True


def get_name_quantity(ws, row):
    #task_table.append(["###", "###"])
    name_col = 5
    quantity_col = 6

    add_to_table(ws, row, name_col)
    add_to_table(ws, row, quantity_col)

    if newer_Excel(ws):
        gatunek_col = 19
    else:
        gatunek_col = 18

    if ws.cell(row, gatunek_col).value is not None:
        add_to_table(ws, row, gatunek_col)
        print(gatunek_col)

def get_dimenstions(ws, row):
    if newer_Excel(ws):
        a_col = 22
        b_col = 23
        c_col = 24
        diameter_col = 20
        length_col = 21
    else:
        a_col = 21
        b_col = 22
        c_col = 23
        diameter_col = 19
        length_col = 20

    a = ws.cell(row, a_col)
    b = ws.cell(row, b_col)
    c = ws.cell(row, c_col)

    diameter = ws.cell(row, diameter_col)
    length = ws.cell(row, length_col)

    if all(cell.value is not None for cell in (a, b, c)):
        add_to_table(ws, row, a_col)
        add_to_table(ws, row, b_col)
        add_to_table(ws, row, c_col)

    elif all(cell.value is not None for cell in (diameter,length)):
        add_to_table(ws, row, diameter_col)
        add_to_table(ws, row, length_col)


def get_line(target_id):                                        # find lines with index
    counter = 0
    for sheet in reversed(get_month_sheet()):
        ws = wb[sheet]

        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=4).value
            value = str(cell_value).strip()
            if str(value) == target_id:
                if check_if_canceled(ws, row):
                    continue
                newer_Excel(ws)
                get_name_quantity(ws, row)
                get_dimenstions(ws, row)
                look_for_operation(ws, row)
                print(task_table)
                part_2.main(target_id, task_table)
                task_table.clear()
                counter += 1
        if counter == 3:
            break



def main(index):
    get_line(index)

